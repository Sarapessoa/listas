import re
import openpyxl

# Função para extrair as informações de uma linha no arquivo M3U
def extract_info(line):
    duration_match = re.search(r'#EXTINF:(-?\d+)', line)
    tvg_id_match = re.search(r'tvg-id="([^"]+)"', line)
    tvg_name_match = re.search(r'tvg-name="([^"]+)"', line)
    tvg_logo_match = re.search(r'tvg-logo="([^"]+)"', line)
    group_title_match = re.search(r'group-title="([^"]+)"', line)
    url = line.strip()

    duration = int(duration_match.group(1)) if duration_match else None
    tvg_id = tvg_id_match.group(1) if tvg_id_match else None
    tvg_name = tvg_name_match.group(1) if tvg_name_match else None
    tvg_logo = tvg_logo_match.group(1) if tvg_logo_match else None
    group_title = group_title_match.group(1) if group_title_match else None

    return {
        'Duration': duration,
        'Tag ID': tvg_id,
        'Name': tvg_name,
        'Logo': tvg_logo,
        'Group': group_title,
        'URL': url,
    }

# Função para ler o arquivo M3U e extrair as informações
def parse_m3u_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    info_list = []
    current_info = None

    for line in lines:
        if line.startswith('#EXTINF:'):
            if current_info:
                info_list.append(current_info)
            current_info = extract_info(line)
        elif current_info:
            current_info['URL'] = line.strip()

    if current_info:
        info_list.append(current_info)

    return info_list

num = '02'

file_path = f'../playlist_{num}/playlist_{num}.m3u'
excel_file_path = f'../playlist_{num}/playlist_{num}.xlsx'

playlist_info = parse_m3u_file(file_path)

# Criar um arquivo Excel
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Playlist Info'

# Definir os cabeçalhos da planilha
headers = ['Duration', 'Tag ID', 'Name', 'Logo', 'Group', 'URL']
for col_num, header in enumerate(headers, 1):
    worksheet.cell(row=1, column=col_num, value=header)

# Preencher a planilha com os dados
for row_num, info in enumerate(playlist_info, 2):
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=row_num, column=col_num, value=info.get(header, ''))

# Salvar o arquivo Excel
workbook.save(excel_file_path)

print(f'Planilha criada e salva em {excel_file_path}')
