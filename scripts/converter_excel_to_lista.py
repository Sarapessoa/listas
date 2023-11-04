import openpyxl
from openpyxl.styles import PatternFill

# Função para criar um arquivo M3U com base nos dados da planilha
def create_m3u_file(excel_file_path, m3u_file_path):
    # Carregar a planilha do Excel
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active

    # Inicializar o conteúdo do arquivo M3U
    m3u_content = "#EXTM3U\n"
    m3u_content += "#EXT-X-SESSION-DATA:DATA-ID=\"com.xui.1_5_5r2\"\n"

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    row_counter = 2

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        duration, tvg_id, tvg_name, tvg_logo, group, url = row

        cell = worksheet.cell(row=row_counter, column=1)

        if cell.fill.start_color.index != red_fill.start_color.index:

            m3u_content += f"#EXTINF:{duration} tvg-id=\"{tvg_id}\" tvg-name=\"{tvg_name}\" tvg-logo=\"{tvg_logo}\" group-title=\"{group}\",{tvg_name}\n"
            m3u_content += f"{url}\n"

        row_counter += 1 

    # Salvar o conteúdo no arquivo M3U
    with open(m3u_file_path, 'w', encoding='utf-8') as m3u_file:
        m3u_file.write(m3u_content)

    print(f'Arquivo M3U criado e salvo em {m3u_file_path}')

num = '03'
versao = '1.0'

excel_file_path = f"../playlist_{num}/playlist_{num}.xlsx"
m3u_file_path = f"../playlist_{num}/versoes/playlist_{versao}.m3u"
create_m3u_file(excel_file_path, m3u_file_path)
